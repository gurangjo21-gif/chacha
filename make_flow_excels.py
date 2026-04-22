"""
흐름 단위 의견수렴 엑셀 생성기 v2 (계층 구조)
- 10개 FLOW_GROUP 각각에 .xlsx 파일 생성
- 각 파일에: 안내 + 흐름개요 + BU 종합시트 + 화면별 시트
- BU 시트: 4블록 의견만 (업무 전체에 대한 의견)
- 화면 시트: 캔버스 1개 + 4블록 의견 (그 화면 전용)
- 시트 메타데이터는 A1 (row 1 숨김) 에 저장 → merge 스크립트가 읽음
"""
import os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HTML_PATH = r'c:\20260421_ideamoa\index.html'
OUT_DIR   = r'c:\20260421_ideamoa\의견수렴_팀별'
os.makedirs(OUT_DIR, exist_ok=True)

# ───────────────────── Parse JS data from index.html ─────────────────────
with open(HTML_PATH, 'r', encoding='utf-8') as f:
    html = f.read()

def extract_js_array(name, text):
    m = re.search(rf'const {name}\s*=\s*\[', text)
    if not m: return ''
    start = m.end() - 1; depth = 0; end = None
    for i in range(start, len(text)):
        c = text[i]
        if c == '[': depth += 1
        elif c == ']':
            depth -= 1
            if depth == 0: end = i+1; break
    return text[start:end]

def parse_js_objects(arr_text):
    objs = []
    depth = 0; start = None; in_str = False; quote = None; i = 0
    while i < len(arr_text):
        c = arr_text[i]
        if in_str:
            if c == '\\': i += 2; continue
            if c == quote: in_str = False
        else:
            if c in "'\"": in_str = True; quote = c
            elif c == '{':
                if depth == 0: start = i
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0 and start is not None:
                    objs.append(arr_text[start:i+1]); start = None
        i += 1
    return objs

def js_obj_to_dict(s):
    d = {}
    body = s.strip()[1:-1]
    pairs = []
    depth = 0; in_str = False; quote = None; buf = ''
    for c in body:
        if in_str:
            buf += c
            if c == quote and (len(buf)<2 or buf[-2]!='\\'): in_str = False
            continue
        if c in "'\"": in_str = True; quote = c; buf += c; continue
        if c in '[{': depth += 1
        elif c in ']}': depth -= 1
        if c == ',' and depth == 0:
            pairs.append(buf.strip()); buf = ''
        else:
            buf += c
    if buf.strip(): pairs.append(buf.strip())
    for p in pairs:
        m = re.match(r'(\w+)\s*:\s*(.+)$', p, re.DOTALL)
        if not m: continue
        k, v = m.group(1), m.group(2).strip()
        d[k] = parse_value(v)
    return d

def parse_value(v):
    v = v.strip()
    if v.startswith('['): return parse_array(v)
    if v.startswith("'") and v.endswith("'"): return v[1:-1]
    if v.startswith('"') and v.endswith('"'): return v[1:-1]
    if v.startswith('{'): return js_obj_to_dict(v)
    if v == 'true':  return True
    if v == 'false': return False
    if v == 'null':  return None
    return v

def parse_array(s):
    s = s.strip()[1:-1]
    items = []
    depth = 0; in_str = False; quote = None; buf = ''
    for c in s:
        if in_str:
            buf += c
            if c == quote and (len(buf)<2 or buf[-2]!='\\'): in_str = False
            continue
        if c in "'\"": in_str = True; quote = c; buf += c; continue
        if c in '[{': depth += 1
        elif c in ']}': depth -= 1
        if c == ',' and depth == 0:
            if buf.strip(): items.append(parse_value(buf.strip()))
            buf = ''
        else:
            buf += c
    if buf.strip(): items.append(parse_value(buf.strip()))
    return items

BUS    = [js_obj_to_dict(o) for o in parse_js_objects(extract_js_array('BUS', html))]
EDGES  = [js_obj_to_dict(o) for o in parse_js_objects(extract_js_array('BU_EDGES', html))]
GROUPS = [js_obj_to_dict(o) for o in parse_js_objects(extract_js_array('FLOW_GROUPS', html))]
print(f'BUS={len(BUS)}, EDGES={len(EDGES)}, GROUPS={len(GROUPS)}')

# ───────────────────── Styles ─────────────────────
FNT_TITLE   = Font(name='맑은 고딕', size=15, bold=True, color='FFFFFF')
FNT_TITLE_S = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
FNT_H1      = Font(name='맑은 고딕', size=13, bold=True, color='111827')
FNT_H2      = Font(name='맑은 고딕', size=11, bold=True, color='18181b')
FNT_LBL     = Font(name='맑은 고딕', size=10, bold=True, color='3f3f46')
FNT_SUB     = Font(name='맑은 고딕', size=9,  color='71717a')
FNT_SUB_S   = Font(name='맑은 고딕', size=8,  color='a1a1aa')
FNT_TXT     = Font(name='맑은 고딕', size=10, color='18181b')
FNT_MONO    = Font(name='Consolas',  size=9,  color='3f3f46')
FNT_LINK    = Font(name='맑은 고딕', size=10, color='2563eb', underline='single')
FNT_WHITE_B = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')

FILL_DARK   = PatternFill('solid', fgColor='111827')
FILL_BU     = PatternFill('solid', fgColor='1e40af')
FILL_SCREEN = PatternFill('solid', fgColor='334155')
FILL_ACCENT = PatternFill('solid', fgColor='2563eb')
FILL_HEAD   = PatternFill('solid', fgColor='f3f4f6')
FILL_LOCK   = PatternFill('solid', fgColor='f7f7f8')
FILL_KEEP   = PatternFill('solid', fgColor='e8faf3')
FILL_IMP    = PatternFill('solid', fgColor='fff3e9')
FILL_DEL    = PatternFill('solid', fgColor='fee8e8')
FILL_NEW    = PatternFill('solid', fgColor='f3eefe')
FILL_BANNER = PatternFill('solid', fgColor='fef3c7')
FILL_GRID_A = PatternFill('solid', fgColor='ffffff')
FILL_GRID_B = PatternFill('solid', fgColor='fbfbfc')
FILL_FLOW_STEP = PatternFill('solid', fgColor='ffedd5')
FILL_ROLE_LIST  = PatternFill('solid', fgColor='dbeafe')  # list
FILL_ROLE_MODE  = PatternFill('solid', fgColor='ecfccb')  # mode
FILL_ROLE_PROC  = PatternFill('solid', fgColor='fed7aa')  # proc
FILL_ROLE_POPUP = PatternFill('solid', fgColor='e9d5ff')  # popup
FILL_ROLE_OTHER = PatternFill('solid', fgColor='e5e7eb')  # other

thin = Side(style='thin', color='e5e7eb')
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
dash = Side(style='dashed', color='cbd5e1')
BORDER_DASH = Border(left=dash, right=dash, top=dash, bottom=dash)

ALIGN_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_LT = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

CANVAS_ROWS = 20
CANVAS_COLS = 18

ROLE_META = {
    'list':  {'ic':'📋', 'lbl':'목록', 'fill':FILL_ROLE_LIST},
    'mode':  {'ic':'✏️', 'lbl':'입력', 'fill':FILL_ROLE_MODE},
    'proc':  {'ic':'⚡', 'lbl':'처리', 'fill':FILL_ROLE_PROC},
    'popup': {'ic':'🔎', 'lbl':'팝업', 'fill':FILL_ROLE_POPUP},
    'other': {'ic':'🧩', 'lbl':'기타', 'fill':FILL_ROLE_OTHER},
}

def classify_screen(filename):
    n = filename.lower()
    if '_proc' in n or '_release' in n: return 'proc'
    if n.startswith('popup_') or '_popup' in n: return 'popup'
    if '_list' in n: return 'list'
    if '_mode' in n or '_setting' in n: return 'mode'
    if '_detail' in n: return 'list'
    return 'other'

# ───────────────────── Sheet name utilities ─────────────────────
# Excel 시트명 제약: 31자, :\/?*[] 불가
# 중복 방지 위해 각 wb 별 사용중인 이름 추적

def sanitize(s):
    for ch in ':\\/?*[]':
        s = s.replace(ch, '_')
    return s

def unique_name(used, base, prefix='', max_len=31):
    """중복 방지. 기본이름 + (필요 시 suffix). 31자 제한."""
    full = prefix + base
    if len(full) > max_len:
        full = full[:max_len]
    full = sanitize(full)
    if full not in used:
        used.add(full); return full
    # 중복 시 _2, _3 붙여가며
    i = 2
    while True:
        tail = f'_{i}'
        truncated = (prefix + base)[:max_len - len(tail)]
        candidate = sanitize(truncated + tail)
        if candidate not in used:
            used.add(candidate); return candidate
        i += 1

def screen_sheet_name(used, filename):
    """화면 파일명 → 시트명. e.g. popup_serial_return_release2_proc.html → popup_ser_ret_rel2_proc"""
    base = filename.replace('.html', '')
    return unique_name(used, base, prefix='')

def bu_sheet_name(used, bu_id):
    """BU 종합 시트명. e.g. receipt-main → BU_receipt_main"""
    return unique_name(used, bu_id.replace('-','_'), prefix='BU_', max_len=31)

# ───────────────────── Helpers ─────────────────────
def neighbors(buid):
    ups = [e for e in EDGES if e.get('to') == buid]
    dns = [e for e in EDGES if e.get('from') == buid]
    return ups, dns

def bu_by_id(bid):
    return next((b for b in BUS if b.get('id')==bid), None)

def bu_display(bid):
    b = bu_by_id(bid)
    return f"{b.get('ic','')} {b.get('name','')}" if b else bid

def add_meta_row(ws, meta_dict, last_col_letter):
    """시트 A1 에 meta 문자열 저장 후 row 1 숨김. 파싱 포맷: META|key=val|key=val"""
    parts = [f'{k}={v}' for k, v in meta_dict.items()]
    ws['A1'] = 'META|' + '|'.join(parts)
    ws['A1'].font = Font(size=8, color='cccccc')
    ws.row_dimensions[1].hidden = True

# ───────────────────── Canvas builder (공통) ─────────────────────
def draw_canvas(ws, start_row, cols=CANVAS_COLS, rows=CANVAS_ROWS):
    """캔버스 영역 (점선 격자)"""
    for rr in range(start_row, start_row + rows):
        ws.row_dimensions[rr].height = 22
        for cc in range(2, 2 + cols):
            cell = ws.cell(row=rr, column=cc)
            cell.border = BORDER_DASH
            cell.fill = FILL_GRID_A if (rr+cc)%2==0 else FILL_GRID_B
    return start_row + rows

def draw_opinion_blocks(ws, start_row, cols=CANVAS_COLS):
    """의견 4블록 (2x2). cols 는 전체 가로 너비."""
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=1+cols)
    c = ws.cell(row=start_row, column=2, value='💬 의견 4블록 — 번호(①②③)로 짚어주세요')
    c.font = FNT_H2; c.alignment = ALIGN_L
    ws.row_dimensions[start_row].height = 22
    r = start_row + 1

    half = cols // 2
    blocks = [
        ('🟢 유지 · 이건 꼭 남겨주세요',     FILL_KEEP, '059669'),
        ('🟡 개선 · 이렇게 바뀌면 좋겠어요', FILL_IMP,  'c2410c'),
        ('🔴 삭제 · 이건 안 써요',           FILL_DEL,  'dc2626'),
        ('💡 신규 · 이게 있으면 좋겠어요',   FILL_NEW,  '7c3aed'),
    ]
    BODY_ROWS = 6
    for i, (title, fill, accent) in enumerate(blocks):
        row_offset = (i // 2) * (1 + BODY_ROWS + 1)  # title + body + gap
        col_offset = (i % 2) * half
        start_col = 2 + col_offset
        end_col = 2 + col_offset + half - 1
        t_row = r + row_offset
        b_start = t_row + 1; b_end = t_row + BODY_ROWS
        ws.merge_cells(start_row=t_row, start_column=start_col, end_row=t_row, end_column=end_col)
        c = ws.cell(row=t_row, column=start_col, value=title)
        c.font = Font(name='맑은 고딕', size=11, bold=True, color=accent)
        c.fill = fill; c.alignment = ALIGN_L
        ws.row_dimensions[t_row].height = 22
        for cc in range(start_col, end_col+1):
            ws.cell(row=t_row, column=cc).border = BORDER_ALL
        for br in range(b_start, b_end+1):
            ws.merge_cells(start_row=br, start_column=start_col, end_row=br, end_column=end_col)
            c = ws.cell(row=br, column=start_col)
            c.font = FNT_TXT; c.alignment = ALIGN_LT
            c.fill = PatternFill('solid', fgColor='ffffff')
            ws.row_dimensions[br].height = 22
            for col in range(start_col, end_col+1):
                ws.cell(row=br, column=col).border = BORDER_ALL
    end_row = start_row + 1 + 2 * (1 + BODY_ROWS + 1) - 1
    return end_row + 1

def draw_footer(ws, row, cols=CANVAS_COLS):
    """작성자 / 작성일"""
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value='작성자'); c.font = FNT_LBL; c.fill = FILL_HEAD; c.alignment = ALIGN_L
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
    ws.cell(row=row, column=4).border = BORDER_ALL; ws.cell(row=row, column=4).fill = PatternFill('solid', fgColor='ffffff')
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    c = ws.cell(row=row, column=8, value='작성일'); c.font = FNT_LBL; c.fill = FILL_HEAD; c.alignment = ALIGN_L
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=1+cols)
    ws.cell(row=row, column=10).border = BORDER_ALL; ws.cell(row=row, column=10).fill = PatternFill('solid', fgColor='ffffff')
    for col in range(2, 2+cols):
        ws.cell(row=row, column=col).border = BORDER_ALL
    ws.row_dimensions[row].height = 26

def init_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    for col_idx in range(2, 2 + CANVAS_COLS):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
    ws.column_dimensions[get_column_letter(2 + CANVAS_COLS)].width = 2

# ───────────────────── BU Summary Sheet ─────────────────────
def add_bu_sheet(wb, bu, step_idx, group, used_names):
    sn = bu_sheet_name(used_names, bu['id'])
    ws = wb.create_sheet(sn)
    init_sheet(ws)
    last_col = get_column_letter(1 + CANVAS_COLS)
    add_meta_row(ws, {'type':'bu', 'bu':bu['id']}, last_col)

    # Title bar (진한 파란색 - BU 전용)
    ws.merge_cells(f'B2:{last_col}3')
    step_prefix = f"[STEP {step_idx:02d}] " if step_idx else '[BU 종합] '
    ws['B2'] = f'{step_prefix}{bu.get("ic","")} {bu.get("name","")}   ·   {bu.get("id","")}'
    ws['B2'].font = FNT_TITLE; ws['B2'].fill = FILL_BU
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

    ws.merge_cells(f'B4:{last_col}4')
    ws['B4'] = f'💼 업무 전체(BU)에 대한 의견을 받는 곳입니다. 이 업무의 개별 화면별 의견은 아래 탭들에 있어요.'
    ws['B4'].font = FNT_SUB; ws['B4'].alignment = ALIGN_L
    ws.row_dimensions[4].height = 22

    r = 6
    # 설명
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+CANVAS_COLS)
    ws.cell(row=r, column=2, value=bu.get('desc','')).font = FNT_TXT
    ws.cell(row=r, column=2).fill = FILL_LOCK
    ws.cell(row=r, column=2).alignment = ALIGN_L
    for col in range(2, 2+CANVAS_COLS): ws.cell(row=r, column=col).border = BORDER_ALL
    ws.row_dimensions[r].height = 30
    r += 1

    # 현재 화면 역할별 분류
    asis = bu.get('asis',[]) or []
    by_role = {'list':[],'mode':[],'proc':[],'popup':[],'other':[]}
    for f in asis: by_role[classify_screen(f)].append(f)

    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value='현재 화면 구성').font = FNT_LBL
    ws.cell(row=r, column=2).fill = FILL_HEAD; ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=1+CANVAS_COLS)
    parts = []
    for role in ['list','mode','proc','popup','other']:
        if by_role[role]:
            m = ROLE_META[role]
            parts.append(f"{m['ic']}{m['lbl']} {len(by_role[role])}")
    ws.cell(row=r, column=4, value=f'총 {len(asis)}개 화면 · ' + ' · '.join(parts)).font = FNT_TXT
    ws.cell(row=r, column=4).fill = FILL_LOCK; ws.cell(row=r, column=4).alignment = ALIGN_L
    for col in range(2, 2+CANVAS_COLS): ws.cell(row=r, column=col).border = BORDER_ALL
    ws.row_dimensions[r].height = 24
    r += 1

    # 연관 테이블
    tables = bu.get('tables',[]) or []
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value='연관 테이블').font = FNT_LBL
    ws.cell(row=r, column=2).fill = FILL_HEAD; ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=1+CANVAS_COLS)
    ws.cell(row=r, column=4, value=', '.join(f'🗄 {t}' for t in tables) if tables else '—').font = FNT_TXT
    ws.cell(row=r, column=4).fill = FILL_LOCK; ws.cell(row=r, column=4).alignment = ALIGN_L
    for col in range(2, 2+CANVAS_COLS): ws.cell(row=r, column=col).border = BORDER_ALL
    ws.row_dimensions[r].height = 22
    r += 1

    # 상류/하류
    ups, dns = neighbors(bu['id'])
    up_txt = ' · '.join(bu_display(e['from']) for e in ups) if ups else '— (시작 업무)'
    dn_txt = ' · '.join(bu_display(e['to'])   for e in dns) if dns else '— (종결 업무)'
    for label, txt, fill_color in [('← 상류', up_txt, 'eff4ff'), ('하류 →', dn_txt, 'fff3e9')]:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=2, value=label); c.font = FNT_LBL
        c.fill = PatternFill('solid', fgColor=fill_color); c.alignment = ALIGN_L
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=1+CANVAS_COLS)
        c = ws.cell(row=r, column=4, value=txt); c.font = FNT_TXT; c.fill = FILL_LOCK; c.alignment = ALIGN_L
        for col in range(2, 2+CANVAS_COLS): ws.cell(row=r, column=col).border = BORDER_ALL
        ws.row_dimensions[r].height = 24
        r += 1
    r += 1

    # 안내 배너
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+CANVAS_COLS)
    c = ws.cell(row=r, column=2, value='💬 이 "업무 전체"에 대한 의견을 남겨주세요. 특정 화면이 아닌 업무 흐름·전반적인 부분.')
    c.font = FNT_H2; c.fill = FILL_BANNER; c.alignment = ALIGN_L
    ws.row_dimensions[r].height = 24
    r += 1

    # 의견 4블록
    r = draw_opinion_blocks(ws, r, cols=CANVAS_COLS)
    r += 1
    draw_footer(ws, r)

    # 이 BU 의 화면 목록 (참고)
    r += 3
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+CANVAS_COLS)
    ws.cell(row=r, column=2, value=f'📂 이 업무의 {len(asis)}개 화면 (각 화면은 별도 시트로 있어요)').font = FNT_H2
    ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.row_dimensions[r].height = 22
    r += 1
    hdr = ['#', '화면 파일명', '역할', '비고']
    widths = [4, 24, 12, 20]
    for i, h in enumerate(hdr):
        cell = ws.cell(row=r, column=2+i, value=h)
        cell.font = FNT_WHITE_B; cell.fill = FILL_ACCENT; cell.alignment = ALIGN_C
        cell.border = BORDER_ALL
    r += 1
    for i, f in enumerate(asis, 1):
        role = classify_screen(f)
        m = ROLE_META[role]
        c1 = ws.cell(row=r, column=2, value=i); c1.font = FNT_TXT; c1.alignment = ALIGN_C; c1.border = BORDER_ALL
        c2 = ws.cell(row=r, column=3, value=f); c2.font = FNT_MONO; c2.alignment = ALIGN_L; c2.border = BORDER_ALL
        c3 = ws.cell(row=r, column=4, value=f"{m['ic']} {m['lbl']}"); c3.font = FNT_TXT
        c3.fill = m['fill']; c3.alignment = ALIGN_C; c3.border = BORDER_ALL
        c4 = ws.cell(row=r, column=5, value=''); c4.border = BORDER_ALL
        r += 1

    return sn

# ───────────────────── Screen Sheet ─────────────────────
def add_screen_sheet(wb, bu, filename, seq_in_bu, total_in_bu, used_names):
    sn = screen_sheet_name(used_names, filename)
    ws = wb.create_sheet(sn)
    init_sheet(ws)
    last_col = get_column_letter(1 + CANVAS_COLS)
    role = classify_screen(filename)
    m = ROLE_META[role]

    add_meta_row(ws, {'type':'screen', 'bu':bu['id'], 'screen':filename, 'role':role}, last_col)

    # Title bar
    ws.merge_cells(f'B2:{last_col}2')
    ws['B2'] = f'{m["ic"]} {filename}     ({seq_in_bu}/{total_in_bu} — {bu.get("ic","")} {bu.get("name","")})'
    ws['B2'].font = FNT_TITLE; ws['B2'].fill = FILL_SCREEN
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 26

    # Sub
    ws.merge_cells(f'B3:{last_col}3')
    ws['B3'] = f'역할: {m["ic"]} {m["lbl"]} · 소속 업무: {bu.get("id","")} · 업무 설명: {bu.get("desc","")}'
    ws['B3'].font = FNT_SUB; ws['B3'].alignment = ALIGN_L
    ws.row_dimensions[3].height = 20

    # Canvas banner
    r = 5
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+CANVAS_COLS)
    c = ws.cell(row=r, column=2, value=f'🎨 [{filename}] 이 화면이 이렇게 되면 좋겠다 — 자유롭게 그려주세요 (셀 병합·색칠·글쓰기·도형)')
    c.font = FNT_H2; c.fill = FILL_BANNER; c.alignment = ALIGN_L
    ws.row_dimensions[r].height = 24
    r += 1

    # Canvas
    r = draw_canvas(ws, r)
    r += 1

    # Opinion 4 blocks
    r = draw_opinion_blocks(ws, r, cols=CANVAS_COLS)

    # Footer
    draw_footer(ws, r)
    return sn

# ───────────────────── Guide Sheet ─────────────────────
def add_guide_sheet(wb, group, total_bus, total_screens):
    ws = wb.create_sheet('00_안내', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    for col, w in zip('BCDEFGHIJ', [22, 22, 22, 22, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:J3')
    c = ws['B2']; c.value = f'{group.get("label","")}  의견수집'
    c.font = FNT_TITLE; c.fill = FILL_DARK
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 30

    ws.merge_cells('B4:J4')
    ws['B4'] = f'이 파일에는 {total_bus}개 업무와 그에 속한 {total_screens}개 화면이 담겨 있습니다. 각 화면마다 그릴 수 있는 캔버스 1개씩 있어요.'
    ws['B4'].font = FNT_SUB; ws['B4'].alignment = ALIGN_L
    ws.row_dimensions[4].height = 30

    guide = [
        ('1️⃣ 흐름개요',    '[01_흐름개요] 시트에서 이 파일이 다루는 업무와 화면 전체를 봅니다. BU 이름 클릭하면 해당 업무로 이동.'),
        ('2️⃣ BU 종합 의견', '파란색 [BU_xxx] 시트: 그 "업무 전체" 에 대한 의견을 남기는 곳. 흐름·전반 관련.'),
        ('3️⃣ 화면별 캔버스','회색 [receipt], [popup_svc] 같은 시트: 그 화면을 어떻게 바꾸고 싶은지 직접 그리고 의견 남기기.'),
        ('4️⃣ 의견 4블록',   '각 시트 하단의 🟢유지 🟡개선 🔴삭제 💡신규 — 번호(①②③) 로 짚어서 적어주세요.'),
        ('5️⃣ 저장·제출',   'Ctrl+S 로 저장 후 담당자에게 이메일/메신저. 파일명에 본인 이름을 붙여주세요 (예: 홍길동_01_일반CS흐름.xlsx)'),
    ]
    r = 6
    for t, d in guide:
        ws.merge_cells(f'B{r}:C{r}')
        c = ws[f'B{r}']; c.value = t; c.font = FNT_H2; c.fill = FILL_HEAD; c.alignment = ALIGN_L
        ws.merge_cells(f'D{r}:J{r}')
        c = ws[f'D{r}']; c.value = d; c.font = FNT_TXT; c.alignment = ALIGN_L
        for col in 'BCDEFGHIJ': ws[f'{col}{r}'].border = BORDER_ALL
        ws.row_dimensions[r].height = 32
        r += 1

    r += 1
    ws.merge_cells(f'B{r}:J{r}'); c = ws[f'B{r}']; c.value = '📌 색상 범례'; c.font = FNT_H1
    r += 1
    legends = [('🟢 유지', FILL_KEEP, '이건 꼭 남겨주세요'),
               ('🟡 개선', FILL_IMP,  '이렇게 바뀌면 좋겠어요'),
               ('🔴 삭제', FILL_DEL,  '이건 안 써요'),
               ('💡 신규', FILL_NEW,  '새로 있으면 좋겠어요')]
    for t, fill, d in legends:
        ws.merge_cells(f'B{r}:C{r}')
        c = ws[f'B{r}']; c.value = t; c.font = FNT_H2; c.fill = fill; c.alignment = ALIGN_C
        ws.merge_cells(f'D{r}:J{r}')
        c = ws[f'D{r}']; c.value = d; c.font = FNT_TXT; c.alignment = ALIGN_L
        for col in 'BCDEFGHIJ': ws[f'{col}{r}'].border = BORDER_ALL
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    ws.merge_cells(f'B{r}:J{r}')
    ws[f'B{r}'] = f'🌐 참고 사이트 · 브라우저로 보면: https://gurangjo21-gif.github.io/chacha/'
    ws[f'B{r}'].font = FNT_SUB; ws[f'B{r}'].alignment = ALIGN_L

# ───────────────────── Overview Sheet ─────────────────────
def add_overview_sheet(wb, group, bu_sheet_map, screen_sheet_map):
    """bu_sheet_map: {bu_id: sheet_name}, screen_sheet_map: {(bu_id,filename): sheet_name}"""
    ws = wb.create_sheet('01_흐름개요', 1)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    widths = {'B':4, 'C':12, 'D':26, 'E':40, 'F':8, 'G':30}
    for k, v in widths.items(): ws.column_dimensions[k].width = v

    is_flow = group.get('kind') == 'flow'

    ws.merge_cells('B2:G2')
    title = group.get("label","") + (' · 여정 스텝' if is_flow else ' · 포함 업무')
    c = ws['B2']; c.value = title
    c.font = FNT_TITLE; c.fill = FILL_DARK
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 32

    ws.merge_cells('B3:G3')
    if is_flow:
        hint = '업무 이름을 클릭하면 [BU 종합 의견] 시트로, 화면 파일명을 클릭하면 [그 화면 캔버스] 시트로 이동합니다.'
    else:
        hint = '업무 이름 클릭 → [BU 종합 의견] 시트, 화면 파일명 클릭 → [그 화면 캔버스] 시트로 이동.'
    ws['B3'] = hint; ws['B3'].font = FNT_SUB

    r = 5
    for step_idx, bu_id in enumerate(group.get('bus',[]), 1):
        bu = bu_by_id(bu_id)
        if not bu: continue
        bu_sn = bu_sheet_map.get(bu_id, '')
        asis = bu.get('asis',[]) or []

        # BU 헤더 행
        step_label = f'STEP {step_idx:02d}' if is_flow else f'{step_idx}.'
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=f'{step_label}   {bu.get("ic","")} {bu.get("name","")}   ·   {len(asis)}개 화면')
        c.font = FNT_TITLE_S; c.fill = FILL_BU
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        if bu_sn:
            c.hyperlink = f"#'{bu_sn}'!A1"
        ws.row_dimensions[r].height = 26
        r += 1

        # 테이블 헤더
        hdr = ['#', '역할', '화면 파일', '설명', 'BU', 'BU 의견 시트']
        for i, h in enumerate(hdr):
            cell = ws.cell(row=r, column=2+i, value=h)
            cell.font = FNT_LBL; cell.fill = FILL_HEAD; cell.alignment = ALIGN_C
            cell.border = BORDER_ALL
        ws.row_dimensions[r].height = 22
        r += 1

        # 화면 목록
        for i, f in enumerate(asis, 1):
            role = classify_screen(f); m = ROLE_META[role]
            sc_sn = screen_sheet_map.get((bu_id, f), '')
            c1 = ws.cell(row=r, column=2, value=i); c1.font = FNT_TXT; c1.alignment = ALIGN_C; c1.border = BORDER_ALL
            c2 = ws.cell(row=r, column=3, value=f"{m['ic']} {m['lbl']}"); c2.font = FNT_TXT
            c2.fill = m['fill']; c2.alignment = ALIGN_C; c2.border = BORDER_ALL
            c3 = ws.cell(row=r, column=4, value=f)
            c3.font = FNT_LINK; c3.alignment = ALIGN_L; c3.border = BORDER_ALL
            if sc_sn: c3.hyperlink = f"#'{sc_sn}'!A1"
            c4 = ws.cell(row=r, column=5, value=bu.get('desc','')[:50]); c4.font = FNT_SUB; c4.alignment = ALIGN_L; c4.border = BORDER_ALL
            c5 = ws.cell(row=r, column=6, value=bu.get('id','')); c5.font = FNT_MONO; c5.alignment = ALIGN_C; c5.border = BORDER_ALL
            c6 = ws.cell(row=r, column=7, value='→ BU 종합 의견')
            c6.font = FNT_LINK; c6.alignment = ALIGN_C; c6.border = BORDER_ALL
            if bu_sn: c6.hyperlink = f"#'{bu_sn}'!A1"
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1  # gap between BUs

    ws.freeze_panes = 'B5'

# ───────────────────── Build one workbook per group ─────────────────────
def build_flow_workbook(group, idx):
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    bu_sheet_map = {}
    screen_sheet_map = {}

    # Compute totals
    total_screens = 0
    for bu_id in group.get('bus',[]):
        bu = bu_by_id(bu_id)
        if bu: total_screens += len(bu.get('asis',[]) or [])
    total_bus = len(group.get('bus',[]))

    # Add guide + overview placeholders (created first to fix order)
    add_guide_sheet(wb, group, total_bus, total_screens)
    # Overview will be added at position 1 after we know sheet names
    # Build BU + screen sheets first
    is_flow = group.get('kind') == 'flow'
    for step_idx, bu_id in enumerate(group.get('bus',[]), 1):
        bu = bu_by_id(bu_id)
        if not bu: continue
        si = step_idx if is_flow else None
        bu_sn = add_bu_sheet(wb, bu, si, group, used_names)
        bu_sheet_map[bu_id] = bu_sn
        asis = bu.get('asis',[]) or []
        total_in_bu = len(asis)
        for i, f in enumerate(asis, 1):
            sn = add_screen_sheet(wb, bu, f, i, total_in_bu, used_names)
            screen_sheet_map[(bu_id, f)] = sn

    # Now add overview at position 1
    add_overview_sheet(wb, group, bu_sheet_map, screen_sheet_map)

    # Filename
    label_clean = re.sub(r'[^\w가-힣]', '', group.get('label',''))
    suffix = f'_{total_screens}화면' if total_screens else ''
    fname = f'{idx:02d}_{label_clean}_{total_bus}BU{suffix}.xlsx'
    path = os.path.join(OUT_DIR, fname)
    wb.save(path)
    return fname, total_screens

# ───────────────────── Run ─────────────────────
# 기존 파일 정리 (중복 방지)
for old in os.listdir(OUT_DIR):
    if old.endswith('.xlsx'):
        try: os.remove(os.path.join(OUT_DIR, old))
        except: pass

print('')
print('==== 흐름 단위 엑셀 생성 (계층 구조) ====')
ordered = sorted(GROUPS, key=lambda g: int(g.get('order','99')))
created = []
total_screens_all = 0
for i, group in enumerate(ordered, 1):
    fname, n_screens = build_flow_workbook(group, i)
    bu_cnt = len(group.get('bus',[]))
    kind = '[FLOW]' if group.get('kind')=='flow' else '[SUP] '
    try:
        print(f'  [{i:02d}] {kind}  {fname}  ({bu_cnt} BU, {n_screens} 화면)')
    except UnicodeEncodeError:
        print(f'  [{i:02d}] {kind}  ({bu_cnt} BU, {n_screens} screens)')
    created.append(fname)
    total_screens_all += n_screens

print('')
print(f'완료: {len(created)}개 파일 · 총 {total_screens_all}개 화면 캔버스')
print(f'저장 위치: {OUT_DIR}')
