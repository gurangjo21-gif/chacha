"""
KOCOM CS 의견수집 엑셀 생성기
- 45개 업무단위별로 시트 생성
- 각 시트: 상단(BU 정보) + 중단(빈 설계 캔버스) + 하단(의견 4블록)
"""
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName

# ───────────────────── Extract BUS, BU_EDGES, SCENARIOS from index.html ─────────────────────
with open(r'c:\20260421_ideamoa\index.html', 'r', encoding='utf-8') as f:
    html = f.read()

def extract_js_array(name, text):
    m = re.search(rf'const {name}\s*=\s*\[', text)
    if not m:
        return []
    start = m.end() - 1  # position of [
    depth = 0
    end = None
    for i in range(start, len(text)):
        c = text[i]
        if c == '[': depth += 1
        elif c == ']':
            depth -= 1
            if depth == 0:
                end = i + 1
                break
    return text[start:end]

bus_str = extract_js_array('BUS', html)
edges_str = extract_js_array('BU_EDGES', html)
scen_str = extract_js_array('SCENARIOS', html)

# Parse each BU object manually (JS → Python). Split by top-level `},\n{` inside the array.
def parse_js_objects(arr_text):
    """Parse a JS array text into list of dicts by finding balanced {} blocks."""
    objs = []
    depth = 0
    start = None
    in_str = False
    quote = None
    i = 0
    while i < len(arr_text):
        c = arr_text[i]
        if in_str:
            if c == '\\':
                i += 2
                continue
            if c == quote:
                in_str = False
        else:
            if c in "'\"":
                in_str = True
                quote = c
            elif c == '{':
                if depth == 0:
                    start = i
                depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0 and start is not None:
                    objs.append(arr_text[start:i+1])
                    start = None
        i += 1
    return objs

def js_obj_to_dict(s):
    """Naively parse a JS object literal string into a Python dict.
    Supports: string keys/values, arrays of strings, nested simple objects.
    """
    # Normalize: wrap keys in quotes
    # keys look like: word:
    # We'll evaluate as JSON after minor rewrites. Safer: write a tiny parser.
    d = {}
    # Strip outer braces
    body = s.strip()[1:-1]
    # Tokenize by finding "key: value" pairs at depth 0
    pairs = []
    depth = 0
    in_str = False
    quote = None
    buf = ''
    for c in body:
        if in_str:
            buf += c
            if c == quote and (len(buf) < 2 or buf[-2] != '\\'):
                in_str = False
            continue
        if c in "'\"":
            in_str = True
            quote = c
            buf += c
            continue
        if c in '[{':
            depth += 1
        elif c in ']}':
            depth -= 1
        if c == ',' and depth == 0:
            pairs.append(buf.strip())
            buf = ''
        else:
            buf += c
    if buf.strip():
        pairs.append(buf.strip())
    for p in pairs:
        m = re.match(r'(\w+)\s*:\s*(.+)$', p, re.DOTALL)
        if not m:
            continue
        k, v = m.group(1), m.group(2).strip()
        d[k] = parse_value(v)
    return d

def parse_value(v):
    v = v.strip()
    if v.startswith('['):
        return parse_array(v)
    if v.startswith("'") and v.endswith("'"):
        return v[1:-1].replace("\\'", "'").replace('\\n', '\n')
    if v.startswith('"') and v.endswith('"'):
        return v[1:-1]
    if v.startswith('{'):
        return js_obj_to_dict(v)
    if v == 'true':  return True
    if v == 'false': return False
    if v == 'null':  return None
    return v

def parse_array(s):
    """Parse JS array string into python list."""
    s = s.strip()[1:-1]  # strip []
    items = []
    depth = 0
    in_str = False
    quote = None
    buf = ''
    for c in s:
        if in_str:
            buf += c
            if c == quote and (len(buf)<2 or buf[-2] != '\\'):
                in_str = False
            continue
        if c in "'\"":
            in_str = True
            quote = c
            buf += c
            continue
        if c in '[{':
            depth += 1
        elif c in ']}':
            depth -= 1
        if c == ',' and depth == 0:
            if buf.strip():
                items.append(parse_value(buf.strip()))
            buf = ''
        else:
            buf += c
    if buf.strip():
        items.append(parse_value(buf.strip()))
    return items

bus_objs = parse_js_objects(bus_str)
BUS = [js_obj_to_dict(o) for o in bus_objs]
print(f'Parsed {len(BUS)} BUs')

edges_objs = parse_js_objects(edges_str)
EDGES = [js_obj_to_dict(o) for o in edges_objs]
print(f'Parsed {len(EDGES)} edges')

scen_objs = parse_js_objects(scen_str)
SCENARIOS = [js_obj_to_dict(o) for o in scen_objs]
print(f'Parsed {len(SCENARIOS)} scenarios')

DOMAINS = {
  'admin':   '🛠 관리자·마스터',
  'receipt': '📞 수신접수',
  'svc':     '🧑‍🔧 서비스 처리',
  'store':   '📦 자재(일반)',
  'tdb':     '⚡ 뇌전(TDB)',
  'serial':  '🔖 시리얼',
  'forum':   '💬 게시판·소통',
  'resolve': '🗓 연차 처리',
  'cti':     '☎ CTI 콜센터',
  'sms':     '📨 SMS',
  'common':  '🔎 공용 팝업',
  'layout':  '🪟 레이아웃',
  'ajax':    '🔌 AJAX 라우터',
  'util':    '🧰 유틸·기타',
}

# ───────────────────── Build workbook ─────────────────────
wb = Workbook()

# Style helpers
FNT_TITLE   = Font(name='맑은 고딕', size=18, bold=True, color='FFFFFF')
FNT_H1      = Font(name='맑은 고딕', size=14, bold=True, color='111827')
FNT_H2      = Font(name='맑은 고딕', size=12, bold=True, color='18181b')
FNT_LBL     = Font(name='맑은 고딕', size=10, bold=True, color='3f3f46')
FNT_SUB     = Font(name='맑은 고딕', size=9,  color='71717a')
FNT_TXT     = Font(name='맑은 고딕', size=10, color='18181b')
FNT_MONO    = Font(name='Consolas',  size=9,  color='3f3f46')
FNT_MUTED   = Font(name='맑은 고딕', size=9,  color='a1a1aa')
FNT_WHITE_B = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')

FILL_DARK   = PatternFill('solid', fgColor='111827')
FILL_ACCENT = PatternFill('solid', fgColor='2563eb')
FILL_HEAD   = PatternFill('solid', fgColor='f3f4f6')
FILL_CANVAS = PatternFill('solid', fgColor='fafafa')
FILL_LOCK   = PatternFill('solid', fgColor='f7f7f8')
FILL_KEEP   = PatternFill('solid', fgColor='e8faf3')
FILL_IMP    = PatternFill('solid', fgColor='fff3e9')
FILL_DEL    = PatternFill('solid', fgColor='fee8e8')
FILL_NEW    = PatternFill('solid', fgColor='f3eefe')
FILL_GRID_A = PatternFill('solid', fgColor='ffffff')
FILL_GRID_B = PatternFill('solid', fgColor='fbfbfc')

thin = Side(style='thin', color='e5e7eb')
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
dash = Side(style='dashed', color='cbd5e1')
BORDER_DASH = Border(left=dash, right=dash, top=dash, bottom=dash)

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_LT = Alignment(horizontal='left',  vertical='top',    wrap_text=True)

# ───────────────────── Sheet 1: 안내 ─────────────────────
ws = wb.active
ws.title = '00_안내'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
for col, w in zip('BCDEFGHIJ', [24, 24, 24, 24, 14, 14, 14, 14, 14]):
    ws.column_dimensions[col].width = w

ws.merge_cells('B2:J3')
ws['B2'] = 'KOCOM CS 차세대 전산 · 의견수집 엑셀북'
ws['B2'].font = FNT_TITLE
ws['B2'].fill = FILL_DARK
ws['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

ws.merge_cells('B4:J4')
ws['B4'] = '각 업무단위(BU) 시트에 당신이 원하는 화면을 직접 그려주세요. 셀을 색칠하고, 글씨를 쓰고, 선을 그어도 됩니다.'
ws['B4'].font = FNT_SUB
ws['B4'].alignment = ALIGN_L

# 범례 — 사용법 3단계
guide = [
    ('1️⃣ 시트 선택', '아래쪽 탭에서 내가 아는 업무를 찾아 클릭합니다. (예: "수신접수_메인", "자재_신청")'),
    ('2️⃣ 상단 확인', '시트 상단의 회색 박스(잠금)에는 현재 업무 정보가 있습니다. 무엇을 다루는 업무인지 참고만 하세요.'),
    ('3️⃣ 캔버스에 그리기', '가운데 흰색 격자는 빈 캔버스입니다. 셀을 합치고(Alt+H,M,M) 색칠하고 글을 써서 원하는 화면을 그려주세요.'),
    ('4️⃣ 의견 4칸 작성', '아래쪽 초록/노랑/빨강/보라 4칸에 각각 "유지/개선/삭제/신규" 의견을 번호(①②③) 로 적어주세요.'),
    ('5️⃣ 저장 후 전달', 'Ctrl+S로 저장 후 파일을 담당자에게 이메일이나 메신저로 보내주세요. 여러 분의 의견이 합쳐집니다.'),
]
r = 6
for title, desc in guide:
    ws.merge_cells(f'B{r}:C{r}')
    ws[f'B{r}'] = title
    ws[f'B{r}'].font = FNT_H2
    ws[f'B{r}'].fill = FILL_HEAD
    ws[f'B{r}'].alignment = ALIGN_L
    ws.merge_cells(f'D{r}:J{r}')
    ws[f'D{r}'] = desc
    ws[f'D{r}'].font = FNT_TXT
    ws[f'D{r}'].alignment = ALIGN_L
    for col in 'BCDEFGHIJ':
        ws[f'{col}{r}'].border = BORDER_ALL
    ws.row_dimensions[r].height = 28
    r += 1

# 범례 — 의견 4블록 색상
r += 1
ws.merge_cells(f'B{r}:J{r}')
ws[f'B{r}'] = '📌 의견 4블록 색상 범례'
ws[f'B{r}'].font = FNT_H1
r += 1
legends = [
    ('🟢 유지', FILL_KEEP,  '이건 꼭 남겨주세요 · 지금이 편함'),
    ('🟡 개선', FILL_IMP,   '이렇게 바뀌면 더 좋겠어요'),
    ('🔴 삭제', FILL_DEL,   '이건 안 써요 · 자리만 차지'),
    ('💡 신규', FILL_NEW,   '이게 새로 있으면 좋겠어요'),
]
for title, fill, desc in legends:
    ws.merge_cells(f'B{r}:C{r}')
    ws[f'B{r}'] = title
    ws[f'B{r}'].font = FNT_H2
    ws[f'B{r}'].fill = fill
    ws[f'B{r}'].alignment = ALIGN_C
    ws.merge_cells(f'D{r}:J{r}')
    ws[f'D{r}'] = desc
    ws[f'D{r}'].font = FNT_TXT
    ws[f'D{r}'].alignment = ALIGN_L
    for col in 'BCDEFGHIJ':
        ws[f'{col}{r}'].border = BORDER_ALL
    ws.row_dimensions[r].height = 24
    r += 1

# 그리기 팁
r += 1
ws.merge_cells(f'B{r}:J{r}')
ws[f'B{r}'] = '🎨 캔버스 그리기 팁'
ws[f'B{r}'].font = FNT_H1
r += 1
tips = [
    '여러 셀 선택 후 셀 병합 (리본 → 홈 → 병합하고 가운데 맞춤)',
    '셀 배경색으로 영역 구분 (리본 → 홈 → 채우기 색)',
    '굵은 테두리로 박스 그리기 (리본 → 홈 → 테두리)',
    '삽입 → 도형 으로 화살표·말풍선 추가 가능',
    '잘못 그렸으면 Ctrl+Z 로 되돌리기',
    '완성 후 Ctrl+S 로 저장 → 담당자에게 전달',
]
for t in tips:
    ws.merge_cells(f'B{r}:J{r}')
    ws[f'B{r}'] = '  • ' + t
    ws[f'B{r}'].font = FNT_TXT
    ws[f'B{r}'].alignment = ALIGN_L
    r += 1

# ───────────────────── Sheet 2: 전체 BU 목록 ─────────────────────
ws = wb.create_sheet('01_업무단위_전체')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
widths = {'B':4, 'C':18, 'D':11, 'E':24, 'F':42, 'G':8, 'H':8, 'I':30, 'J':30}
for k,v in widths.items():
    ws.column_dimensions[k].width = v

ws.merge_cells('B2:J2')
ws['B2'] = '🗺 45개 업무단위 전체 지도'
ws['B2'].font = FNT_TITLE
ws['B2'].fill = FILL_DARK
ws['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[2].height = 40

ws.merge_cells('B3:J3')
ws['B3'] = '각 행은 한 개의 업무단위. 시트 탭을 클릭하거나 [시트명] 열의 이름을 참고해 해당 시트로 이동하세요.'
ws['B3'].font = FNT_SUB

hdr_row = 5
headers = ['#','도메인','BU ID','업무 이름','설명','화면수','TO-BE','상류 (들어옴)','하류 (나감)']
for i, h in enumerate(headers):
    cell = ws.cell(row=hdr_row, column=i+2, value=h)
    cell.font = FNT_WHITE_B
    cell.fill = FILL_ACCENT
    cell.alignment = ALIGN_C
    cell.border = BORDER_ALL
ws.row_dimensions[hdr_row].height = 28

# Sort BUs by domain order then name
DOM_ORDER = list(DOMAINS.keys())
def dom_idx(b): return DOM_ORDER.index(b.get('dom','util'))
bus_sorted = sorted(BUS, key=lambda b: (dom_idx(b), b.get('id','')))

# Compute upstream/downstream per BU
def neighbors(buid):
    ups = [e for e in EDGES if e.get('to') == buid]
    dns = [e for e in EDGES if e.get('from') == buid]
    return ups, dns

def bu_name_by_id(bid):
    for b in BUS:
        if b.get('id') == bid:
            return f"{b.get('ic','')} {b.get('name','')}"
    return bid

row = hdr_row + 1
bu_sheet_names = {}  # bu_id -> sheet name (for hyperlinks)
for idx, b in enumerate(bus_sorted, 1):
    ic = b.get('ic','')
    bid = b.get('id','')
    name = b.get('name','')
    desc = b.get('desc','')
    dom = DOMAINS.get(b.get('dom','util'), b.get('dom','util'))
    ups, dns = neighbors(bid)
    up_txt = ' · '.join(bu_name_by_id(e['from']) for e in ups) if ups else '—'
    dn_txt = ' · '.join(bu_name_by_id(e['to'])   for e in dns) if dns else '—'
    asis_cnt = len(b.get('asis', []) or [])
    tobe_cnt = len(b.get('tobe', []) or [])
    tobe_mark = f'✅ {tobe_cnt}' if tobe_cnt else '⚠ 미정'

    values = [idx, dom, bid, f"{ic} {name}", desc, asis_cnt, tobe_mark, up_txt, dn_txt]
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=i+2, value=v)
        c.font = FNT_TXT if i != 2 else FNT_MONO
        c.alignment = ALIGN_LT if i in (4,7,8) else ALIGN_C if i in (0,5,6) else ALIGN_L
        c.border = BORDER_ALL
        if tobe_cnt == 0 and i == 6:
            c.fill = FILL_IMP
        if tobe_cnt > 0 and i == 6:
            c.fill = FILL_KEEP
    ws.row_dimensions[row].height = max(34, 18 + max(len(up_txt), len(dn_txt))//40 * 14)
    row += 1

# Freeze header
ws.freeze_panes = ws.cell(row=hdr_row+1, column=2)

# ───────────────────── Sheet 3: 여정맵 ─────────────────────
ws = wb.create_sheet('02_여정맵')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
for col_letter, w in [('B',4), ('C',22), ('D',22), ('E',22), ('F',22), ('G',22), ('H',22), ('I',22), ('J',22), ('K',22), ('L',22)]:
    ws.column_dimensions[col_letter].width = w

ws.merge_cells('B2:L2')
ws['B2'] = '🔗 4가지 대표 여정 · 업무는 흐름이다'
ws['B2'].font = FNT_TITLE
ws['B2'].fill = FILL_DARK
ws['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[2].height = 40

ws.merge_cells('B3:L3')
ws['B3'] = 'KOCOM CS의 모든 업무는 아래 4개 여정 중 하나로 귀결됩니다. 각 스텝은 하나의 업무단위 시트에 해당합니다.'
ws['B3'].font = FNT_SUB

scen_colors = {'c-blue':'eff4ff', 'c-pink':'fdf0f7', 'c-violet':'f3eefe', 'c-amber':'fff3e9'}
scen_accent = {'c-blue':'2563eb', 'c-pink':'db2777', 'c-violet':'7c3aed', 'c-amber':'c2410c'}

r = 5
for sc in SCENARIOS:
    color_key = sc.get('color', 'c-blue')
    bg = scen_colors.get(color_key, 'eff4ff')
    ac = scen_accent.get(color_key, '2563eb')
    # Header
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    ws.cell(row=r, column=2, value=f"{sc.get('ic','')} {sc.get('name','')}")
    ws.cell(row=r, column=2).font = Font(name='맑은 고딕', size=13, bold=True, color=ac)
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=bg)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[r].height = 28
    r += 1
    # Description
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    ws.cell(row=r, column=2, value=sc.get('desc',''))
    ws.cell(row=r, column=2).font = FNT_SUB
    ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.row_dimensions[r].height = 20
    r += 1
    # Steps row (up to 11 steps)
    steps = sc.get('steps', [])
    for i, st in enumerate(steps[:11]):
        buid = st.get('bu','')
        b = next((x for x in BUS if x.get('id')==buid), None)
        if not b: continue
        col = 2 + i
        cell = ws.cell(row=r, column=col, value=f"STEP {i+1}\n{b.get('ic','')} {b.get('name','')}")
        cell.font = Font(name='맑은 고딕', size=10, bold=True, color='18181b')
        cell.fill = PatternFill('solid', fgColor='ffffff')
        cell.border = Border(
            left=Side(style='medium', color=ac),
            right=Side(style='medium', color=ac),
            top=Side(style='medium', color=ac),
            bottom=Side(style='medium', color=ac),
        )
        cell.alignment = ALIGN_C
    ws.row_dimensions[r].height = 40
    r += 1
    # Action
    for i, st in enumerate(steps[:11]):
        col = 2 + i
        cell = ws.cell(row=r, column=col, value=st.get('act',''))
        cell.font = FNT_SUB
        cell.alignment = ALIGN_C
    ws.row_dimensions[r].height = 30
    r += 2  # gap

# ───────────────────── Per-BU sheets ─────────────────────
def safe_sheet_name(s):
    # Excel limits: 31 chars, no :\/?*[]
    for ch in ':\\/?*[]':
        s = s.replace(ch, '_')
    return s[:31]

# Canvas size
CANVAS_ROWS = 22
CANVAS_COLS = 18  # columns B..S

for b in bus_sorted:
    bid = b.get('id','')
    name = b.get('name','')
    ic = b.get('ic','')
    desc = b.get('desc','')
    dom = DOMAINS.get(b.get('dom','util'), b.get('dom','util'))
    tables = b.get('tables', []) or []
    asis_files = b.get('asis', []) or []
    tobe_files = b.get('tobe', []) or []

    raw = bid.replace('-','_')
    sheet_name = safe_sheet_name(raw)
    ws = wb.create_sheet(sheet_name)
    bu_sheet_names[bid] = sheet_name
    ws.sheet_view.showGridLines = False
    # column widths: A narrow, B..S moderate, T right narrow
    ws.column_dimensions['A'].width = 2
    for col_idx in range(2, 2 + CANVAS_COLS):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
    ws.column_dimensions[get_column_letter(2 + CANVAS_COLS)].width = 2

    # ── Title bar ──
    last_col_letter = get_column_letter(1 + CANVAS_COLS)
    ws.merge_cells(f'B2:{last_col_letter}3')
    ws['B2'] = f'{ic} {name}   ·   {bid}'
    ws['B2'].font = FNT_TITLE
    ws['B2'].fill = FILL_DARK
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

    ws.merge_cells(f'B4:{last_col_letter}4')
    ws['B4'] = desc
    ws['B4'].font = FNT_SUB
    ws['B4'].alignment = ALIGN_L
    ws.row_dimensions[4].height = 22

    # ── Info grid: domain / tables / asis count / tobe count ──
    info_rows = [
        ('도메인', dom),
        ('현재 화면', f'{len(asis_files)}개 파일 · ' + ', '.join(asis_files[:4]) + (' …' if len(asis_files) > 4 else '')),
        ('연관 테이블', ', '.join(f'🗄 {t}' for t in tables) if tables else '—'),
        ('새 버전 초안', f'✅ {len(tobe_files)}개 있음' if tobe_files else '⚠ 미정 — 이 시트의 당신 그림이 곧 초안이 됩니다'),
    ]
    r = 6
    for lbl, val in info_rows:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2, value=lbl).font = FNT_LBL
        ws.cell(row=r, column=2).fill = FILL_HEAD
        ws.cell(row=r, column=2).alignment = ALIGN_L
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=1+CANVAS_COLS)
        c = ws.cell(row=r, column=4, value=val)
        c.font = FNT_TXT
        c.alignment = ALIGN_L
        c.fill = FILL_LOCK
        for col in range(2, 2+CANVAS_COLS):
            ws.cell(row=r, column=col).border = BORDER_ALL
        ws.row_dimensions[r].height = 22
        r += 1

    # ── Upstream / Downstream ──
    ups, dns = neighbors(bid)
    up_txt = ' · '.join(bu_name_by_id(e['from']) for e in ups) if ups else '— (상류 없음)'
    dn_txt = ' · '.join(bu_name_by_id(e['to'])   for e in dns) if dns else '— (하류 없음)'
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value='← 상류 (들어옴)').font = FNT_LBL
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor='eff4ff')
    ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=1+CANVAS_COLS)
    ws.cell(row=r, column=4, value=up_txt).font = FNT_TXT
    ws.cell(row=r, column=4).fill = FILL_LOCK
    ws.cell(row=r, column=4).alignment = ALIGN_L
    for col in range(2, 2+CANVAS_COLS):
        ws.cell(row=r, column=col).border = BORDER_ALL
    ws.row_dimensions[r].height = 26
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value='하류 (나감) →').font = FNT_LBL
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor='fff3e9')
    ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=1+CANVAS_COLS)
    ws.cell(row=r, column=4, value=dn_txt).font = FNT_TXT
    ws.cell(row=r, column=4).fill = FILL_LOCK
    ws.cell(row=r, column=4).alignment = ALIGN_L
    for col in range(2, 2+CANVAS_COLS):
        ws.cell(row=r, column=col).border = BORDER_ALL
    ws.row_dimensions[r].height = 26
    r += 2  # spacer

    # ── Canvas banner ──
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+CANVAS_COLS)
    ws.cell(row=r, column=2, value='🎨 내가 원하는 화면을 여기에 그려주세요 (셀 병합 · 색칠 · 글쓰기 · 도형 삽입 자유롭게)').font = FNT_H2
    ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor='fef3c7')
    ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.row_dimensions[r].height = 24
    r += 1

    # ── Canvas grid ──
    canvas_start = r
    canvas_end = r + CANVAS_ROWS - 1
    for rr in range(canvas_start, canvas_end + 1):
        ws.row_dimensions[rr].height = 22
        for cc in range(2, 2 + CANVAS_COLS):
            cell = ws.cell(row=rr, column=cc)
            cell.border = BORDER_DASH
            cell.fill = FILL_GRID_A if (rr+cc) % 2 == 0 else FILL_GRID_B
    r = canvas_end + 2

    # ── Feedback 4 blocks ──
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+CANVAS_COLS)
    ws.cell(row=r, column=2, value='💬 이 업무에 대한 의견 4블록 (번호로 적어주세요. 예: "③번을 크게" "⑦번 없앴으면")').font = FNT_H2
    ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.row_dimensions[r].height = 22
    r += 1

    half = CANVAS_COLS // 2  # 9
    # First row: 유지 / 개선
    # Second row: 삭제 / 신규
    blocks = [
        ('🟢 유지 · 이건 꼭 남겨주세요', FILL_KEEP, '059669'),
        ('🟡 개선 · 이렇게 바뀌면 좋겠어요', FILL_IMP, 'c2410c'),
        ('🔴 삭제 · 이건 안 써요',         FILL_DEL, 'dc2626'),
        ('💡 신규 · 이게 있으면 좋겠어요', FILL_NEW, '7c3aed'),
    ]
    for i, (title, fill, accent) in enumerate(blocks):
        row_offset = (i // 2) * 9
        col_offset = (i % 2) * half
        start_col = 2 + col_offset
        end_col = 2 + col_offset + half - 1
        title_row = r + row_offset
        body_start = title_row + 1
        body_end = title_row + 7

        ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=end_col)
        c = ws.cell(row=title_row, column=start_col, value=title)
        c.font = Font(name='맑은 고딕', size=11, bold=True, color=accent)
        c.fill = fill
        c.alignment = ALIGN_L
        ws.row_dimensions[title_row].height = 22
        for cc in range(start_col, end_col+1):
            ws.cell(row=title_row, column=cc).border = BORDER_ALL

        for body_row in range(body_start, body_end+1):
            ws.merge_cells(start_row=body_row, start_column=start_col, end_row=body_row, end_column=end_col)
            cc = ws.cell(row=body_row, column=start_col)
            cc.font = FNT_TXT
            cc.alignment = ALIGN_LT
            cc.fill = PatternFill('solid', fgColor='ffffff')
            ws.row_dimensions[body_row].height = 22
            for col in range(start_col, end_col+1):
                ws.cell(row=body_row, column=col).border = BORDER_ALL
    r += 18  # 2 rows of (1 title + 7 body) = 16 + spacers

    # ── Footer: author / date ──
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value='작성자').font = FNT_LBL
    ws.cell(row=r, column=2).fill = FILL_HEAD
    ws.cell(row=r, column=2).alignment = ALIGN_L
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    ws.cell(row=r, column=4).border = BORDER_ALL
    ws.cell(row=r, column=4).fill = PatternFill('solid', fgColor='ffffff')

    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
    ws.cell(row=r, column=8, value='작성일').font = FNT_LBL
    ws.cell(row=r, column=8).fill = FILL_HEAD
    ws.cell(row=r, column=8).alignment = ALIGN_L
    ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=1+CANVAS_COLS)
    ws.cell(row=r, column=10).border = BORDER_ALL
    ws.cell(row=r, column=10).fill = PatternFill('solid', fgColor='ffffff')
    for col in range(2, 2+CANVAS_COLS):
        ws.cell(row=r, column=col).border = BORDER_ALL
    ws.row_dimensions[r].height = 26

    # Set print area and orientation
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

# Save
out = r'c:\20260421_ideamoa\KOCOM_CS_의견수집.xlsx'
wb.save(out)
print(f'저장 완료: {out}')
print(f'총 시트: {len(wb.sheetnames)}개')
