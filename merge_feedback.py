"""
의견수렴 엑셀 자동 취합기 v2 (BU + 화면 계층 구조 대응)
- 의견수렴_회수/ 폴더의 모든 *.xlsx 를 읽어서
- BU 종합 시트의 4블록 의견 + 화면별 시트의 4블록 의견 추출
- 통합 엑셀 + JSON 생성

사용법:
  1. 직원들이 채워서 보낸 파일을 의견수렴_회수/ 폴더에 넣음
  2. python merge_feedback.py
  3. 결과: 통합_의견.xlsx + 통합_의견.json
"""
import os, re, glob, json
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

IN_DIR  = r'c:\20260421_ideamoa\의견수렴_회수'
OUT_XLSX = r'c:\20260421_ideamoa\통합_의견.xlsx'
OUT_JSON = r'c:\20260421_ideamoa\통합_의견.json'
INDEX_HTML = r'c:\20260421_ideamoa\index.html'

os.makedirs(IN_DIR, exist_ok=True)

# ────────────────────── BU meta from index.html ──────────────────────
def load_bu_meta():
    with open(INDEX_HTML, 'r', encoding='utf-8') as f:
        html = f.read()
    meta = {}
    for m in re.finditer(
        r"\{id:'([a-z][a-z0-9-]+)',dom:'(\w+)',ic:'([^']*)',name:'([^']+)',desc:'([^']+)'",
        html):
        meta[m.group(1)] = {'id':m.group(1),'dom':m.group(2),'ic':m.group(3),'name':m.group(4),'desc':m.group(5)}
    return meta

BU_META = load_bu_meta()
print(f'[i] BU 메타: {len(BU_META)}개 로드')

# ────────────────────── Parse metadata from A1 ──────────────────────
def parse_meta(val):
    """'META|type=screen|bu=receipt-main|screen=receipt.html|role=other' → dict"""
    if not val or not isinstance(val, str): return None
    if not val.startswith('META|'): return None
    d = {}
    for part in val.split('|')[1:]:
        if '=' in part:
            k, v = part.split('=', 1)
            d[k.strip()] = v.strip()
    return d

# ────────────────────── Opinion block extraction ──────────────────────
CAT_MARKERS = {
    'keep': '🟢 유지',
    'imp':  '🟡 개선',
    'del':  '🔴 삭제',
    'new':  '💡 신규',
}
CAT_LABEL = {'keep':'유지','imp':'개선','del':'삭제','new':'신규'}
CAT_EMOJI = {'keep':'🟢','imp':'🟡','del':'🔴','new':'💡'}

def find_block_positions(ws):
    positions = {}
    for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=False):
        for cell in row:
            if not cell.value or not isinstance(cell.value, str): continue
            v = cell.value.strip()
            for cat, marker in CAT_MARKERS.items():
                if v.startswith(marker):
                    positions[cat] = (cell.row, cell.column)
                    break
    return positions

_MARKER_STARTS = tuple(CAT_MARKERS.values())
def read_opinions_below(ws, row, col, max_rows=6):
    """타이틀 아래 N줄 읽기. 다른 블록 타이틀(🟢🟡🔴💡) 만나면 중단."""
    opinions = []
    for r in range(row+1, row+1+max_rows):
        v = ws.cell(row=r, column=col).value
        if v is None: continue
        if isinstance(v, str):
            v = v.strip()
            if not v: continue
            if v.startswith(_MARKER_STARTS): break  # 다음 블록 타이틀 시작
            opinions.append(v)
        else:
            opinions.append(str(v))
    return opinions

def find_footer(ws):
    author = ''; date = ''
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if not cell.value or not isinstance(cell.value, str): continue
            v = cell.value.strip()
            if v == '작성자':
                for offset in (2, 1, 3):
                    nxt = ws.cell(row=cell.row, column=cell.column + offset).value
                    if nxt: author = str(nxt).strip(); break
            elif v == '작성일':
                for offset in (2, 1, 3):
                    nxt = ws.cell(row=cell.row, column=cell.column + offset).value
                    if nxt: date = str(nxt).strip(); break
    return author, date

# ────────────────────── Process single file ──────────────────────
def process_file(xlsx_path):
    results = []
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    except Exception as e:
        print(f'  [!] 로드 실패: {os.path.basename(xlsx_path)} — {e}')
        return results

    src = os.path.basename(xlsx_path)
    for sheet_name in wb.sheetnames:
        # Skip guide/overview
        if sheet_name.startswith(('00_','01_','02_','안내','개요')): continue
        ws = wb[sheet_name]
        meta = parse_meta(ws['A1'].value)
        if not meta: continue  # 메타 없으면 무시 (사용자 직접 추가 시트 등)

        sheet_type = meta.get('type')
        bu_id = meta.get('bu')
        if bu_id not in BU_META: continue

        positions = find_block_positions(ws)
        if not positions: continue

        author, date = find_footer(ws)
        if not author:
            fn_base = os.path.splitext(src)[0]
            parts = fn_base.split('_')
            if parts and parts[0] and not parts[0].isdigit():
                author = parts[0]

        screen = meta.get('screen', '') if sheet_type == 'screen' else ''
        role = meta.get('role', '') if sheet_type == 'screen' else ''
        scope = 'BU 종합' if sheet_type == 'bu' else '화면별'

        for cat, (row, col) in positions.items():
            opinions = read_opinions_below(ws, row, col, max_rows=6)
            for op in opinions:
                results.append({
                    'bu':       bu_id,
                    'scope':    scope,       # 'BU 종합' | '화면별'
                    'screen':   screen,       # 파일명 or ''
                    'role':     role,         # 'list'|'mode'|'proc'|'popup'|'other'|''
                    'category': cat,
                    'text':     op,
                    'author':   author or '익명',
                    'date':     date or '',
                    'source':   src,
                })
    return results

# ────────────────────── Collect ──────────────────────
print(f'[i] 스캔 폴더: {IN_DIR}')
files = sorted(glob.glob(os.path.join(IN_DIR, '*.xlsx')))
files = [f for f in files if not os.path.basename(f).startswith('~$')]
print(f'[i] 읽을 파일: {len(files)}개')

all_ops = []
for fp in files:
    fn = os.path.basename(fp)
    ops = process_file(fp)
    all_ops.extend(ops)
    print(f'  · {fn}: {len(ops)}개 의견')

if not all_ops:
    print('')
    print('[경고] 추출된 의견이 0개입니다.')
    print('       의견수렴_회수/ 에 채워진 엑셀이 있는지 확인하세요.')

# ────────────────────── Group ──────────────────────
by_bu       = defaultdict(lambda: {'bu_overall': defaultdict(list), 'by_screen': defaultdict(lambda: defaultdict(list))})
by_category = defaultdict(list)
by_author   = defaultdict(int)

for o in all_ops:
    if o['scope'] == 'BU 종합':
        by_bu[o['bu']]['bu_overall'][o['category']].append(o)
    else:
        by_bu[o['bu']]['by_screen'][o['screen']][o['category']].append(o)
    by_category[o['category']].append(o)
    by_author[o['author']] += 1

# ────────────────────── Styles (output xlsx) ──────────────────────
FNT_TITLE   = Font(name='맑은 고딕', size=15, bold=True, color='FFFFFF')
FNT_H1      = Font(name='맑은 고딕', size=13, bold=True, color='111827')
FNT_H2      = Font(name='맑은 고딕', size=11, bold=True, color='18181b')
FNT_LBL     = Font(name='맑은 고딕', size=10, bold=True, color='3f3f46')
FNT_SUB     = Font(name='맑은 고딕', size=9,  color='71717a')
FNT_TXT     = Font(name='맑은 고딕', size=10, color='18181b')
FNT_MONO    = Font(name='Consolas',  size=9,  color='3f3f46')
FNT_WHITE_B = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')

FILL_DARK   = PatternFill('solid', fgColor='111827')
FILL_BU     = PatternFill('solid', fgColor='1e40af')
FILL_SCREEN = PatternFill('solid', fgColor='334155')
FILL_ACCENT = PatternFill('solid', fgColor='2563eb')
FILL_HEAD   = PatternFill('solid', fgColor='f3f4f6')
FILL_KEEP   = PatternFill('solid', fgColor='e8faf3')
FILL_IMP    = PatternFill('solid', fgColor='fff3e9')
FILL_DEL    = PatternFill('solid', fgColor='fee8e8')
FILL_NEW    = PatternFill('solid', fgColor='f3eefe')
CAT_FILL = {'keep':FILL_KEEP,'imp':FILL_IMP,'del':FILL_DEL,'new':FILL_NEW}

thin = Side(style='thin', color='e5e7eb')
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

wb_out = Workbook()
wb_out.remove(wb_out.active)

# ── 00_요약 ──
ws = wb_out.create_sheet('00_요약')
ws.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJK', [2, 22, 10, 10, 14, 14, 2, 22, 22, 22, 22]):
    ws.column_dimensions[col].width = w

ws.merge_cells('B2:K3')
ws['B2'] = '📊 의견수렴 취합 결과'
ws['B2'].font = FNT_TITLE; ws['B2'].fill = FILL_DARK
ws['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[2].height = 32

ws.merge_cells('B4:K4')
bu_count = sum(1 for v in by_bu.values() if sum(len(x) for x in v['bu_overall'].values())>0 or v['by_screen'])
screen_count = sum(len(v['by_screen']) for v in by_bu.values())
ws['B4'] = (f'취합: {datetime.now().strftime("%Y-%m-%d %H:%M")} · 파일 {len(files)}개 · '
            f'의견 {len(all_ops)}건 · BU {len(by_bu)}개 · 화면 {screen_count}개 · 참여자 {len(by_author)}명')
ws['B4'].font = FNT_SUB

# 카테고리별
ws['B6'] = '분류별'; ws['B6'].font = FNT_H1
for i, h in enumerate(['분류','의견 수','비중']):
    c = ws.cell(row=7, column=2+i, value=h)
    c.font = FNT_WHITE_B; c.fill = FILL_ACCENT; c.alignment = ALIGN_C; c.border = BORDER_ALL
r = 8
total = len(all_ops) or 1
for cat in ['keep','imp','del','new']:
    n = len(by_category[cat])
    ws.cell(row=r, column=2, value=f'{CAT_EMOJI[cat]} {CAT_LABEL[cat]}').font = FNT_TXT
    ws.cell(row=r, column=2).fill = CAT_FILL[cat]; ws.cell(row=r, column=2).border = BORDER_ALL
    ws.cell(row=r, column=3, value=n).font = FNT_TXT; ws.cell(row=r, column=3).alignment = ALIGN_C
    ws.cell(row=r, column=3).border = BORDER_ALL
    ws.cell(row=r, column=4, value=f'{n*100/total:.1f}%').font = FNT_TXT
    ws.cell(row=r, column=4).alignment = ALIGN_C; ws.cell(row=r, column=4).border = BORDER_ALL
    r += 1

# Scope 별 (BU 종합 vs 화면별)
ws['E6'] = '범위별'; ws['E6'].font = FNT_H1
for i, h in enumerate(['범위','의견 수']):
    c = ws.cell(row=7, column=5+i, value=h)
    c.font = FNT_WHITE_B; c.fill = FILL_ACCENT; c.alignment = ALIGN_C; c.border = BORDER_ALL
by_scope = {'BU 종합':0, '화면별':0}
for o in all_ops: by_scope[o['scope']] += 1
r = 8
for scope, n in by_scope.items():
    ws.cell(row=r, column=5, value=scope).font = FNT_TXT; ws.cell(row=r, column=5).border = BORDER_ALL
    ws.cell(row=r, column=6, value=n).font = FNT_TXT; ws.cell(row=r, column=6).alignment = ALIGN_C
    ws.cell(row=r, column=6).border = BORDER_ALL
    r += 1

# BU별 (Top 15)
ws['H6'] = 'BU별 의견 수 (Top 15)'; ws['H6'].font = FNT_H1
hdr = ['BU 이름','전체','BU 종합','화면별','유지/개선/삭제/신규']
for i, h in enumerate(hdr):
    col = 8 + i
    c = ws.cell(row=7, column=col, value=h)
    c.font = FNT_WHITE_B; c.fill = FILL_ACCENT; c.alignment = ALIGN_C; c.border = BORDER_ALL

bu_counts = []
for bu, data in by_bu.items():
    overall_cnt = sum(len(v) for v in data['bu_overall'].values())
    screen_cnt  = sum(sum(len(v) for v in screens.values()) for screens in data['by_screen'].values())
    cat_cnt = defaultdict(int)
    for v in data['bu_overall'].values():
        for o in v: cat_cnt[o['category']] += 1
    for screens in data['by_screen'].values():
        for v in screens.values():
            for o in v: cat_cnt[o['category']] += 1
    bu_counts.append((bu, overall_cnt + screen_cnt, overall_cnt, screen_cnt, cat_cnt))
bu_counts.sort(key=lambda x: -x[1])

r = 8
for bu, tot, oc, sc, cc in bu_counts[:15]:
    bm = BU_META.get(bu, {})
    label = f"{bm.get('ic','')} {bm.get('name', bu)}"
    ws.cell(row=r, column=8, value=label).font = FNT_TXT; ws.cell(row=r, column=8).border = BORDER_ALL
    ws.cell(row=r, column=9, value=tot).font = Font(bold=True, size=10)
    ws.cell(row=r, column=9).alignment = ALIGN_C; ws.cell(row=r, column=9).border = BORDER_ALL
    ws.cell(row=r, column=10, value=oc).font = FNT_TXT
    ws.cell(row=r, column=10).alignment = ALIGN_C; ws.cell(row=r, column=10).border = BORDER_ALL
    ws.cell(row=r, column=11, value=sc).font = FNT_TXT
    ws.cell(row=r, column=11).alignment = ALIGN_C; ws.cell(row=r, column=11).border = BORDER_ALL
    r += 1

# 참여자
start_r = r + 2
ws.cell(row=start_r, column=2, value='👤 참여자별').font = FNT_H1
r = start_r + 1
for i, h in enumerate(['작성자','의견 수']):
    c = ws.cell(row=r, column=2+i, value=h)
    c.font = FNT_WHITE_B; c.fill = FILL_ACCENT; c.alignment = ALIGN_C; c.border = BORDER_ALL
r += 1
for author, cnt in sorted(by_author.items(), key=lambda x: -x[1]):
    ws.cell(row=r, column=2, value=author).font = FNT_TXT; ws.cell(row=r, column=2).border = BORDER_ALL
    ws.cell(row=r, column=3, value=cnt).font = FNT_TXT
    ws.cell(row=r, column=3).alignment = ALIGN_C; ws.cell(row=r, column=3).border = BORDER_ALL
    r += 1

# ── 01_전체의견 (flat table) ──
ws = wb_out.create_sheet('01_전체의견')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
widths = {'B':4, 'C':22, 'D':10, 'E':22, 'F':6, 'G':50, 'H':14, 'I':14, 'J':28}
for k, v in widths.items(): ws.column_dimensions[k].width = v

ws.merge_cells('B2:J2')
ws['B2'] = '📝 전체 의견 (필터·정렬 가능)'
ws['B2'].font = FNT_TITLE; ws['B2'].fill = FILL_DARK
ws['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[2].height = 30

headers = ['#','업무 (BU)','범위','화면 파일','분류','의견','작성자','날짜','원본 파일']
for i, h in enumerate(headers):
    c = ws.cell(row=4, column=2+i, value=h)
    c.font = FNT_WHITE_B; c.fill = FILL_ACCENT; c.alignment = ALIGN_C; c.border = BORDER_ALL
ws.row_dimensions[4].height = 26

r = 5
for idx, o in enumerate(all_ops, 1):
    bm = BU_META.get(o['bu'], {})
    bu_label = f"{bm.get('ic','')} {bm.get('name', o['bu'])}"
    row_vals = [
        idx, bu_label, o['scope'], o['screen'] or '—',
        f"{CAT_EMOJI[o['category']]} {CAT_LABEL[o['category']]}",
        o['text'], o['author'], o['date'], o['source'],
    ]
    for i, v in enumerate(row_vals):
        c = ws.cell(row=r, column=2+i, value=v)
        c.font = FNT_TXT if i != 8 else FNT_MONO
        c.alignment = ALIGN_L
        c.border = BORDER_ALL
        if i == 4: c.fill = CAT_FILL[o['category']]; c.alignment = ALIGN_C
        if i in (0,): c.alignment = ALIGN_C
    r += 1

ws.freeze_panes = 'B5'
ws.auto_filter.ref = f'B4:J{max(r-1, 4)}'

# ── 02~05: 카테고리별 시트 ──
for idx, cat in enumerate(['keep','imp','del','new'], 2):
    label = f'{idx:02d}_{CAT_LABEL[cat]}'
    ws = wb_out.create_sheet(label)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    for k, v in {'B':4,'C':22,'D':10,'E':22,'F':50,'G':16,'H':14,'I':28}.items():
        ws.column_dimensions[k].width = v

    ws.merge_cells('B2:I2')
    c = ws['B2']; c.value = f'{CAT_EMOJI[cat]} {CAT_LABEL[cat]} 의견 · 총 {len(by_category[cat])}건'
    c.font = FNT_TITLE; c.fill = FILL_DARK
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 30

    hdr = ['#','업무 (BU)','범위','화면 파일','의견','작성자','날짜','원본 파일']
    for i, h in enumerate(hdr):
        c = ws.cell(row=4, column=2+i, value=h)
        c.font = FNT_WHITE_B; c.fill = FILL_ACCENT; c.alignment = ALIGN_C; c.border = BORDER_ALL
    ws.row_dimensions[4].height = 26

    r = 5
    sorted_ops = sorted(by_category[cat], key=lambda o: (o['bu'], o['screen'], o['author']))
    for i, o in enumerate(sorted_ops, 1):
        bm = BU_META.get(o['bu'], {})
        bu_label = f"{bm.get('ic','')} {bm.get('name', o['bu'])}"
        for j, v in enumerate([i, bu_label, o['scope'], o['screen'] or '—', o['text'], o['author'], o['date'], o['source']]):
            c = ws.cell(row=r, column=2+j, value=v)
            c.font = FNT_TXT if j != 7 else FNT_MONO
            c.alignment = ALIGN_L if j == 4 else (ALIGN_C if j == 0 else ALIGN_L)
            c.border = BORDER_ALL
        r += 1
    ws.freeze_panes = 'B5'

# ── BU별 통합 시트 ──
def safe_name(s):
    for ch in ':\\/?*[]':
        s = s.replace(ch, '_')
    return s[:31]

for bu_id, data in sorted(by_bu.items(), key=lambda x: -sum(
    [len(v) for v in x[1]['bu_overall'].values()] +
    [len(v) for screens in x[1]['by_screen'].values() for v in screens.values()]
)):
    bm = BU_META.get(bu_id, {})
    sheet_name = safe_name('BU_' + bu_id.replace('-','_'))
    ws = wb_out.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    for k, v in {'B':4, 'C':22, 'D':50, 'E':16, 'F':14, 'G':28}.items():
        ws.column_dimensions[k].width = v

    ws.merge_cells('B2:G3')
    c = ws['B2']; c.value = f"{bm.get('ic','')} {bm.get('name', bu_id)}  ·  {bu_id}"
    c.font = FNT_TITLE; c.fill = FILL_BU
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 28

    tot_overall = sum(len(v) for v in data['bu_overall'].values())
    tot_screen  = sum(sum(len(v) for v in screens.values()) for screens in data['by_screen'].values())
    ws.merge_cells('B4:G4')
    ws['B4'] = f'{bm.get("desc","")}   ·   BU 종합 {tot_overall}건 · 화면별 {tot_screen}건 · 총 {tot_overall+tot_screen}건'
    ws['B4'].font = FNT_SUB

    r = 6
    # BU 종합 섹션
    if tot_overall > 0:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=f'💼 BU 종합 의견 ({tot_overall}건)')
        c.font = FNT_H2; c.fill = FILL_BU; c.font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        c.alignment = ALIGN_L
        ws.row_dimensions[r].height = 22
        r += 1
        for cat in ['keep','imp','del','new']:
            ops = data['bu_overall'].get(cat, [])
            if not ops: continue
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            c = ws.cell(row=r, column=2, value=f'  {CAT_EMOJI[cat]} {CAT_LABEL[cat]} ({len(ops)}건)')
            c.font = FNT_H2; c.fill = CAT_FILL[cat]; c.alignment = ALIGN_L
            ws.row_dimensions[r].height = 20
            r += 1
            for i, h in enumerate(['#','의견','작성자','날짜','원본 파일']):
                cell = ws.cell(row=r, column=2+i, value=h)
                cell.font = FNT_LBL; cell.fill = FILL_HEAD; cell.alignment = ALIGN_C; cell.border = BORDER_ALL
            r += 1
            for i, o in enumerate(ops, 1):
                for j, v in enumerate([i, o['text'], o['author'], o['date'], o['source']]):
                    c = ws.cell(row=r, column=2+j, value=v)
                    c.font = FNT_TXT if j != 4 else FNT_MONO
                    c.alignment = ALIGN_L if j == 1 else (ALIGN_C if j == 0 else ALIGN_L)
                    c.border = BORDER_ALL
                r += 1
        r += 1

    # 화면별 섹션
    if tot_screen > 0:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(row=r, column=2, value=f'📑 화면별 의견 ({tot_screen}건)')
        c.font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF'); c.fill = FILL_SCREEN
        c.alignment = ALIGN_L
        ws.row_dimensions[r].height = 22
        r += 1
        for screen_file, cats in sorted(data['by_screen'].items()):
            screen_total = sum(len(v) for v in cats.values())
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            c = ws.cell(row=r, column=2, value=f'  📄 {screen_file} ({screen_total}건)')
            c.font = FNT_H2; c.fill = FILL_HEAD; c.alignment = ALIGN_L
            ws.row_dimensions[r].height = 20
            r += 1
            for cat in ['keep','imp','del','new']:
                ops = cats.get(cat, [])
                if not ops: continue
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
                c = ws.cell(row=r, column=2, value=f'    {CAT_EMOJI[cat]} {CAT_LABEL[cat]} ({len(ops)}건)')
                c.font = Font(size=10, bold=True, color='3f3f46'); c.fill = CAT_FILL[cat]
                c.alignment = ALIGN_L
                r += 1
                for i, o in enumerate(ops, 1):
                    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
                    c = ws.cell(row=r, column=2, value=f'      {i}. {o["text"]}')
                    c.font = FNT_TXT; c.alignment = ALIGN_L; c.border = BORDER_ALL
                    ws.cell(row=r, column=5, value=o['author']).font = FNT_TXT
                    ws.cell(row=r, column=5).border = BORDER_ALL; ws.cell(row=r, column=5).alignment = ALIGN_L
                    ws.cell(row=r, column=6, value=o['date']).font = FNT_TXT
                    ws.cell(row=r, column=6).border = BORDER_ALL
                    ws.cell(row=r, column=7, value=o['source']).font = FNT_MONO
                    ws.cell(row=r, column=7).border = BORDER_ALL
                    for col in range(2,8): ws.cell(row=r, column=col).border = BORDER_ALL
                    r += 1
            r += 1
    ws.freeze_panes = 'B6'

# Save
wb_out.save(OUT_XLSX)
print('')
print(f'[완료] 엑셀: {OUT_XLSX}  ({len(wb_out.sheetnames)}시트)')

# ── JSON ──
payload = {
    'meta': {
        'generated_at': datetime.now().isoformat(),
        'source_files': len(files),
        'total_opinions': len(all_ops),
        'participants': list(by_author.keys()),
    },
    'by_category': {cat: [{k:v for k,v in o.items()} for o in ops]
                    for cat, ops in by_category.items()},
    'by_bu': {
        bu: {
            'meta': BU_META.get(bu, {'id':bu}),
            'bu_overall': dict(data['bu_overall']),
            'by_screen': {s: dict(cats) for s, cats in data['by_screen'].items()},
        } for bu, data in by_bu.items()
    },
    'all_opinions': all_ops,
}
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(payload, f, ensure_ascii=False, indent=2)
print(f'[완료] JSON: {OUT_JSON}')
print('')
print('=' * 60)
print(f'· 파일: {len(files)}개  · 의견: {len(all_ops)}건')
print(f'· BU: {len(by_bu)}개  · 화면: {screen_count}개  · 참여자: {len(by_author)}명')
print('=' * 60)
print('')
print('다음 단계:')
print(f'  1. {OUT_XLSX} 열기')
print(f'  2. 패턴 분석 필요하면 {os.path.basename(OUT_JSON)} 를 Claude 에게')
